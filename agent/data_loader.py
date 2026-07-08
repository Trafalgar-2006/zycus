"""
data_loader.py — Loads and normalises project Excel files.

Handles:
- Column name differences between S2P and Project B
- #UNPARSEABLE sentinel values
- NaT / NaN gaps
- String-encoded Duration and Variance fields ("7d", "-3d", "0")
- Known sign-inverted Variance rows (documented anomalies)
"""
from __future__ import annotations
import warnings
warnings.filterwarnings("ignore", category=FutureWarning, module="pandas")

import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
import pandas as pd

from config import SIGN_INVERTED_TASKS

# ── Column aliases ───────────────────────────────────────────────────────────
# Project Plan B has 'Variance2' for the per-task baseline delta;
# S2P has 'Variance' in the same semantic position.
_VARIANCE_ALIAS = {"Variance2": "Variance"}
_REQUIRED_COLS  = [
    "Task Name", "Status", "% Complete", "At Risk?", "Schedule Health",
    "On Hold?", "Critical ?", "Total Float", "Duration",
    "Start Date", "End Date", "Baseline Start", "Baseline Finish",
    "Variance", "RAG", "Area", "Phase/Milestone", "Assigned To",
]


def _parse_days(value) -> Optional[float]:
    """Convert '7d', '-3d', '0', 15.0 → float days.  Returns None on failure."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    s = str(value).strip().lower().replace("d", "").replace(" ", "")
    if s in ("", "nan", "#unparseable"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _working_day_diff(d1: datetime, d2: datetime) -> Optional[int]:
    """Return d1 − d2 in working days (Mon–Fri). Positive = d1 is later."""
    try:
        if pd.isnull(d1) or pd.isnull(d2):
            return None
        if not (isinstance(d1, datetime) and isinstance(d2, datetime)):
            return None
        a, b = d1.date(), d2.date()
        sign = 1 if a >= b else -1
        lo, hi = min(a, b), max(a, b)
        return sign * int(np.busday_count(lo, hi))
    except Exception:
        return None


def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename known aliases and add missing columns as NaN.

    Known column name differences across project files:
    - Project B (UniSan): uses 'Schedule Health' instead of 'RAG'.
      Both encode Red/Yellow/Green task health.
    - Project B: had 'Variance2' (per-task) and 'Variance' (project-level stub).
      We promote Variance2 -> Variance and drop the stub to avoid duplicate cols.
    """
    # Variance alias: promote Variance2 -> Variance, drop project-level stub
    if "Variance2" in df.columns:
        if "Variance" in df.columns:
            df = df.drop(columns=["Variance"])
        df = df.rename(columns={"Variance2": "Variance"})

    # RAG alias: 'Schedule Health' used by Project B
    if "RAG" not in df.columns and "Schedule Health" in df.columns:
        df = df.rename(columns={"Schedule Health": "RAG"})

    for col in _REQUIRED_COLS:
        if col not in df.columns:
            df[col] = np.nan

    # Ensure no duplicate column names remain (safety net)
    df = df.loc[:, ~df.columns.duplicated(keep="first")]
    return df


def _clean_sentinel(df: pd.DataFrame) -> pd.DataFrame:
    """Replace '#UNPARSEABLE' with NaN across the whole frame."""
    return df.replace("#UNPARSEABLE", np.nan).infer_objects(copy=False)


def _add_computed_columns(df: pd.DataFrame, project_name: str) -> pd.DataFrame:
    """Add derived numeric columns used downstream."""
    df = df.copy()

    # ── Variance sign (empirically verified: negative = late) ────────────────
    df["variance_days"] = df["Variance"].apply(_parse_days)

    # Flag the 2 known sign-inverted rows as anomalies (don't silently drop)
    df["variance_anomaly"] = df["Task Name"].isin(SIGN_INVERTED_TASKS)

    # variance_sign: 'late' | 'early' | 'on_time' | None
    def _sign(row):
        if row["variance_anomaly"]:
            return "anomaly"
        v = row["variance_days"]
        if v is None:
            return None
        if v < 0:
            return "late"
        if v > 0:
            return "early"
        return "on_time"

    df["variance_sign"] = df.apply(_sign, axis=1)

    # Numeric encoding of sign for ML: late=-1, on_time=0, early=1, anomaly/None=0
    # This is the RELIABLE signal from Variance — sign only, not magnitude.
    sign_map = {"late": -1, "on_time": 0, "early": 1, "anomaly": 0}
    df["variance_sign_code"] = df["variance_sign"].map(sign_map).fillna(0).astype(float)

    # ── Actual elapsed vs planned (uses real dates, NOT Variance or Duration magnitude) ──
    # NOTE: planned_days (Duration) also cascades for parent/summary rows.
    # Empirically verified: 57% of parent-task Durations don't match either
    # sum OR max of their children — same PM-tool rollup mechanism as Variance.
    # actual_minus_planned is reliable only for leaf tasks (no children).
    # For summary rows it inherits the Duration cascade contamination.
    # Downstream callers (monte_carlo, rag_scorer) should be aware of this.
    df["planned_days"] = df["Duration"].apply(_parse_days)
    df["actual_elapsed_wd"] = df.apply(
        lambda r: _working_day_diff(r["End Date"], r["Start Date"]), axis=1
    )
    df["actual_minus_planned"] = df.apply(
        lambda r: (r["actual_elapsed_wd"] - r["planned_days"])
        if (r["actual_elapsed_wd"] is not None and r["planned_days"] is not None)
        else None,
        axis=1,
    )


    # ── Boolean helpers ──────────────────────────────────────────────────────
    df["is_critical"] = df["Critical ?"].apply(lambda x: bool(x) if x is not None else False)
    df["is_on_hold"]  = df["On Hold?"].apply(
        lambda x: str(x).strip().lower() in ("true", "yes", "1") if x is not None else False
    )
    df["is_at_risk"]  = df["At Risk?"].apply(
        lambda x: str(x).strip().lower() in ("true", "yes", "high", "1") if x is not None else False
    )

    # ── Numeric % complete ───────────────────────────────────────────────────
    df["pct_complete"] = pd.to_numeric(df["% Complete"], errors="coerce").clip(0, 1)

    # ── Numeric total float ──────────────────────────────────────────────────
    df["total_float_days"] = pd.to_numeric(df["Total Float"], errors="coerce")

    # ── Status encoded ───────────────────────────────────────────────────────
    status_map = {
        "completed": 4, "in progress": 3, "not started": 2,
        "on hold": 1, "not applicable": 0,
    }
    df["status_code"] = (
        df["Status"].str.lower().str.strip().map(status_map).fillna(2)
    )

    # ── RAG encoded ──────────────────────────────────────────────────────────
    rag_map = {"green": 0, "yellow": 1, "amber": 1, "red": 2}
    df["rag_code"] = (
        df["RAG"].str.lower().str.strip().map(rag_map)
        if df["RAG"].notna().any()
        else np.nan
    )

    df["_project"] = project_name
    return df


def load_project(filepath: str | Path, project_name: str) -> pd.DataFrame:
    """
    Load a single project Excel file and return a clean task-level DataFrame.

    Parameters
    ----------
    filepath     : path to .xlsx file
    project_name : human-readable label used in reports

    Returns
    -------
    pd.DataFrame with all original columns plus derived columns.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Project file not found: {path.resolve()}")

    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Identify the task-plan sheet (first sheet that is not Summary/Comments)
    sheet_name = next(
        (s for s in wb.sheetnames if s.lower() not in ("summary", "comments")),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{sheet_name}' in {path.name} is empty.")

    df = pd.DataFrame(rows[1:], columns=rows[0])
    df = df.reset_index(drop=True)  # prevent duplicate-label errors

    # ── Load Summary sheet ───────────────────────────────────────────────────
    summary = {}
    if "Summary" in wb.sheetnames:
        ws_sum = wb["Summary"]
        for row in ws_sum.iter_rows(values_only=True):
            if row[0] is not None:
                summary[row[0]] = row[1]

    # ── Load Comments sheet ──────────────────────────────────────────────────
    comments = []
    if "Comments" in wb.sheetnames:
        ws_cmt = wb["Comments"]
        for row in ws_cmt.iter_rows(values_only=True):
            if any(c is not None for c in row):
                comments.append(row)

    df = _normalise_columns(df)
    df = _clean_sentinel(df)
    df = _add_computed_columns(df, project_name)

    # Attach metadata
    df.attrs["summary"]  = summary
    df.attrs["comments"] = comments
    df.attrs["project"]  = project_name
    df.attrs["source"]   = str(path.name)

    return df


def load_all_projects(project_files: dict[str, str]) -> dict[str, pd.DataFrame]:
    """Load all projects defined in config.PROJECT_FILES."""
    return {name: load_project(path, name) for name, path in project_files.items()}
